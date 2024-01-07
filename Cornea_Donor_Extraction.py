import streamlit as st
import pandas as pd
import pdf2image
import pytesseract
import re
import os
from io import BytesIO
from PIL import Image
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Function to extract data from PDF
def extract_data_from_pdf(pdf_path):
    # Convert PDF to images
    images = pdf2image.convert_from_path(pdf_path)
    # Extract text from the images using OCR
    ocr_texts = [pytesseract.image_to_string(img) for img in images]

    # Regular expression patterns for extracting values
    patterns = {
        "Tissue ID": r"Tissue\s?ID[:#]?\s?(?P<value>[\d-]+[\s\w]+)",
        "DIN": r"DIN:\s?(?P<value>W\d{4}\s\d{2}\s\d{6})",
        "Product Code": r"Product Code:\s?(?P<value>V\d{7})",
        "Tissue Type": r"Tissue Type:\s?(?P<value>Cornea)",
        "Donor Age": r"Donor Age:\s?(?P<value>\d+)",
        "Epithelium": r"Epithelium:\s?(?P<value>.+?)(?:\n|Descemet's)",
        "Stroma": r"Stroma:\s?(?P<value>.+?)(?:\n|Endothelium)",
        "Descemet's": r"Descemet's:\s?(?P<value>.+?)(?:\n|Endothelium)",
        "Endothelium": r"Endothelium:\s?(?P<value>.+?)(?:\n|Ocular)",
        "Donor Gender": r"Donor Gender:\s?(?P<value>\w+)",
        "Donor Race": r"Donor Race:\s?(?P<value>\w+)",
        "Primary COD": r"Primary COD:\s?(?P<value>CHF)(?![\w\s])",
        "Date-Time of Death": r"Date-Time of Death:\s?(?P<value>[\d-]+\s[\d:]+)",
        "Date-Time of In Situ": r"Date-Time of In Situ:\s?(?P<value>[\d-]+\s[\d:]+)",
        "Ocular Cooling": r"Ocular Cooling:\s?(?P<value>[\d-]+\s[\d:]+)",
        "Total": r"Total:\s?(?P<value>[\d:]+)",
        "Storage Media": r"Storage Media:\s?(?P<value>[\w-]+)",
        "Media Lot#": r"Media Lot#[:\s]?\s?(?P<value>[\d\w-]+)",
        "Approved Usages": r"Approved Usages:\s?(?P<value>.+?)(?:\n)",
        "Lens Type": r"Lens Type:\s?(?P<value>[\w\s]+)",
        "Testing Facility": r"Testing Facility:\s?(?P<value>VRL Eurofins)(?![\w\s])",
        "Endothelial Cell Density": r"Endothelial Cell Density:\s?(?P<value>\d+)",
        "HBcAb (Total)": r"HBcAb \(Total\):\s?(?P<value>\w+)",
        "HCV Ab": r"HCV Ab:\s?(?P<value>\w+)",
        "HIV-1/HCV/HBV NAT (Ultrio)": r"HIV-1/HCV/HBV NAT \(Ultrio\):\s?(?P<value>\w+)",
        "HBsAg": r"HBsAg:\s?(?P<value>\w+)",
        "HIV 1&2 Ab": r"HIV 1&2 Ab:\s?(?P<value>\w+)",
        "RPR": r"RPR:\s?(?P<value>\w+)",
        "Recent hx": r"Recent hx:\s?(?P<value>.+?)(?:\n\n|\Z)",
        "Sars-Cov-2": r"Sars-Cov-2:\s?(?P<value>[\w\s]+)",
        "Antibodies to Cytomegalovirus (CMV)": r"Antibodies to Cytomegalovirus \(CMV\):\s?(?P<value>[\w\s]+)",
        "Toxoplasma IgG": r"Toxoplasma IgG:\s?(?P<value>[\w\s]+)",
        "EBV - Epstein-Barr (EB) Virus": r"EBV - Epstein-Barr \(EB\) Virus:\s?(?P<value>[\w\s]+)"
    }

     # Extract values using the patterns
    extracted_values = {}
    for field, pattern in patterns.items():
        for text in ocr_texts:
            match = re.search(pattern, text, re.DOTALL)
            if match:
                extracted_values[field] = match.group("value").strip()
                break

    return extracted_values

# Streamlit interface
st.title("Donor Cornea PDF Data Extractor")

# Use Streamlit session state to store the extracted data
if 'all_data' not in st.session_state:
    st.session_state['all_data'] = []

uploaded_files = st.file_uploader("Upload PDF files", type=["pdf"], accept_multiple_files=True)

# Process each uploaded file
if uploaded_files:
    for uploaded_file in uploaded_files:
        # Check if file has already been processed
        if uploaded_file.name not in [data.get('filename') for data in st.session_state['all_data']]:
            with st.spinner(f"Processing {uploaded_file.name}..."):
                with NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
                    temp_file.write(uploaded_file.read())
                data = extract_data_from_pdf(temp_file.name)
                # Add filename to the data for tracking
                data['filename'] = uploaded_file.name
                st.session_state['all_data'].append(data)

# Generate Excel button
if st.session_state['all_data'] and st.button("Generate Excel"):
    df = pd.DataFrame(st.session_state['all_data'])
    # Drop the filename column before creating Excel file
    df.drop(columns=['filename'], inplace=True)

    # Save the DataFrame to a BytesIO buffer
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    # Write the buffer to a temporary file
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(buffer.read())
        tmp_path = tmp.name

    # Load the workbook from the temporary file
    workbook = load_workbook(tmp_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Save the modified workbook to a new BytesIO buffer
    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)

    # Offer the Excel file for download
    st.write("### Download Excel File")
    st.download_button("Click to Download", output_buffer, "data.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Clean up the temporary file
    os.remove(tmp_path)